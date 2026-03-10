"""
utils/report_generator.py
==========================
Generates the attendance report in Excel (.xlsx) and CSV formats.

The Excel report has:
- Sheet 1: Full attendance with formatting (green=present, red=absent)
- Sheet 2: Summary stats (total, present %, emotion breakdown)

Uses openpyxl for formatting — pandas alone can't do the colored cells.
"""

import csv
import os
from datetime import datetime
from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import logging

log = logging.getLogger(__name__)


# ─── Color constants ──────────────────────────────────────────────────────────

PRESENT_FILL  = PatternFill("solid", start_color="D6F5D6")  # light green
ABSENT_FILL   = PatternFill("solid", start_color="FFD6D6")  # light red
HEADER_FILL   = PatternFill("solid", start_color="1F3864")  # dark blue
SUBHEAD_FILL  = PatternFill("solid", start_color="2E75B6")  # medium blue
ALT_ROW_FILL  = PatternFill("solid", start_color="F0F4FF")  # very light blue

HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT    = Font(name="Arial", bold=True, size=14, color="1F3864")
BODY_FONT     = Font(name="Arial", size=10)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")


def generate_excel_report(
    records: Dict[str, dict],
    session_start: datetime,
    session_end: datetime,
    excel_path: str,
    csv_path: str,
):
    """
    Main export function. Generates both Excel and CSV outputs.

    Args:
        records:       Dict from AttendanceLog.records
        session_start: When the session started
        session_end:   When it ended
        excel_path:    Where to save the .xlsx file
        csv_path:      Where to save the .csv file
    """
    wb = Workbook()

    _build_attendance_sheet(wb, records, session_start, session_end)
    _build_summary_sheet(wb, records, session_start, session_end)

    wb.save(excel_path)
    log.info(f"Excel report saved: {excel_path}")

    _export_csv(records, session_start, session_end, csv_path)
    log.info(f"CSV report saved: {csv_path}")


# ─── Sheet 1: Attendance ──────────────────────────────────────────────────────

def _build_attendance_sheet(wb, records, session_start, session_end):
    ws = wb.active
    ws.title = "Attendance"

    date_str = session_start.strftime("%A, %d %B %Y")
    window_str = f"{session_start.strftime('%I:%M %p')} – {session_end.strftime('%I:%M %p')}"

    # ── Title block ───────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = "🎓 STUDENT ATTENDANCE REPORT"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Date: {date_str}   |   Session: {window_str}"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="555555")
    ws["A2"].alignment = CENTER

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    # Spacer row
    ws.row_dimensions[3].height = 6

    # ── Headers ────────────────────────────────────────────────────────────────
    headers = ["#", "Student Name", "Status", "Time In", "Emotion", "Confidence (%)"]
    col_widths = [5, 25, 12, 12, 14, 16]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[4].height = 20

    # ── Data rows ──────────────────────────────────────────────────────────────
    sorted_students = sorted(records.items(), key=lambda x: x[0])

    for row_num, (name, data) in enumerate(sorted_students, 1):
        excel_row = row_num + 4
        is_present = data["status"] == "present"

        row_fill = PRESENT_FILL if is_present else ABSENT_FILL
        if not is_present and row_num % 2 == 0:
            row_fill = PatternFill("solid", start_color="FFE5E5")

        row_data = [
            row_num,
            name.replace("_", " "),
            "✓ Present" if is_present else "✗ Absent",
            data["time_in"].strftime("%H:%M:%S") if data["time_in"] else "—",
            data["emotion"].capitalize() if data["emotion"] != "N/A" else "—",
            data["confidence"] if is_present else "—",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = CENTER if col_idx != 2 else LEFT

        ws.row_dimensions[excel_row].height = 18

    # ── Footer totals ─────────────────────────────────────────────────────────
    total_rows = len(records)
    footer_row = total_rows + 5

    present_count = sum(1 for d in records.values() if d["status"] == "present")
    absent_count = total_rows - present_count
    rate = f"{present_count / total_rows * 100:.1f}%" if total_rows > 0 else "0%"

    ws.merge_cells(f"A{footer_row}:B{footer_row}")
    ws[f"A{footer_row}"] = "TOTALS"
    ws[f"A{footer_row}"].font = Font(name="Arial", bold=True, size=10)
    ws[f"A{footer_row}"].fill = SUBHEAD_FILL
    ws[f"A{footer_row}"].font = Font(name="Arial", bold=True, color="FFFFFF")
    ws[f"A{footer_row}"].alignment = CENTER

    ws[f"C{footer_row}"] = f"Present: {present_count}/{total_rows}"
    ws[f"C{footer_row}"].font = Font(name="Arial", bold=True, color="1A7F1A")
    ws[f"C{footer_row}"].alignment = CENTER

    ws[f"D{footer_row}"] = f"Absent: {absent_count}"
    ws[f"D{footer_row}"].font = Font(name="Arial", bold=True, color="CC0000")
    ws[f"D{footer_row}"].alignment = CENTER

    ws[f"F{footer_row}"] = f"Rate: {rate}"
    ws[f"F{footer_row}"].font = Font(name="Arial", bold=True)
    ws[f"F{footer_row}"].alignment = CENTER

    # Freeze header row
    ws.freeze_panes = "A5"


# ─── Sheet 2: Summary ─────────────────────────────────────────────────────────

def _build_summary_sheet(wb, records, session_start, session_end):
    ws = wb.create_sheet("Summary")

    date_str = session_start.strftime("%d/%m/%Y")
    total = len(records)
    present = sum(1 for d in records.values() if d["status"] == "present")
    absent = total - present
    rate = present / total * 100 if total > 0 else 0

    # Emotion breakdown
    from collections import Counter
    emotions = [d["emotion"] for d in records.values() if d["status"] == "present" and d["emotion"] != "N/A"]
    emotion_counts = Counter(emotions)

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "SESSION SUMMARY"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 26

    # Session info
    info_rows = [
        ("Date", date_str),
        ("Session Start", session_start.strftime("%H:%M:%S")),
        ("Session End", session_end.strftime("%H:%M:%S")),
        ("Duration (min)", round((session_end - session_start).total_seconds() / 60, 1)),
    ]

    for i, (key, val) in enumerate(info_rows, 3):
        ws.cell(row=i, column=1, value=key).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=i, column=2, value=val).font = BODY_FONT

    # Stats
    stats_start = 8
    ws.cell(row=stats_start, column=1, value="Attendance Stats").font = Font(name="Arial", bold=True, size=11)

    stats = [
        ("Total Students", total),
        ("Present", present),
        ("Absent", absent),
        ("Attendance Rate", f"{rate:.1f}%"),
    ]

    for i, (k, v) in enumerate(stats, stats_start + 1):
        ws.cell(row=i, column=1, value=k).font = BODY_FONT
        cell = ws.cell(row=i, column=2, value=v)
        cell.font = BODY_FONT
        if k == "Present":
            cell.font = Font(name="Arial", color="1A7F1A", bold=True)
        elif k == "Absent":
            cell.font = Font(name="Arial", color="CC0000", bold=True)

    # Emotion breakdown
    emo_start = stats_start + len(stats) + 2
    ws.cell(row=emo_start, column=1, value="Emotion Breakdown").font = Font(name="Arial", bold=True, size=11)

    for i, (emotion, count) in enumerate(sorted(emotion_counts.items(), key=lambda x: -x[1]), emo_start + 1):
        ws.cell(row=i, column=1, value=emotion.capitalize()).font = BODY_FONT
        ws.cell(row=i, column=2, value=count).font = BODY_FONT

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18


# ─── CSV export ───────────────────────────────────────────────────────────────

def _export_csv(records, session_start, session_end, csv_path):
    """Exports flat CSV with all attendance data."""
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        # Metadata header rows
        writer.writerow(["Attendance Report"])
        writer.writerow(["Date", session_start.strftime("%Y-%m-%d")])
        writer.writerow(["Session", f"{session_start.strftime('%H:%M:%S')} – {session_end.strftime('%H:%M:%S')}"])
        writer.writerow([])

        # Column headers
        writer.writerow(["Student Name", "Status", "Time In", "Emotion", "Confidence (%)"])

        for name, data in sorted(records.items()):
            writer.writerow([
                name.replace("_", " "),
                data["status"].capitalize(),
                data["time_in"].strftime("%H:%M:%S") if data["time_in"] else "",
                data["emotion"].capitalize(),
                data["confidence"] if data["status"] == "present" else "",
            ])

        # Footer summary
        total = len(records)
        present = sum(1 for d in records.values() if d["status"] == "present")
        writer.writerow([])
        writer.writerow(["Total", total])
        writer.writerow(["Present", present])
        writer.writerow(["Absent", total - present])
        writer.writerow(["Rate (%)", f"{present/total*100:.1f}" if total else "0"])
